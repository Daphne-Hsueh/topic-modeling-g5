import json
import openpyxl
from openpyxl.styles import Font

with open("keyword_categories.json") as f:
    categories = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Keywords"

for col, (cat, keywords) in enumerate(categories.items(), start=1):
    ws.cell(row=1, column=col, value=cat).font = Font(bold=True)
    for row, kw in enumerate(keywords, start=2):
        ws.cell(row=row, column=col, value=kw)
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 30

wb.save("data/processed/keyword_dictionary.xlsx")
print("Saved data/processed/keyword_dictionary.xlsx")